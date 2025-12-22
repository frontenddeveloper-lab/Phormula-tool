from flask import Blueprint, request, jsonify
import jwt
from werkzeug.utils import secure_filename
import pandas as pd
from sqlalchemy import create_engine, Table, MetaData, Column, Integer, String, Float, text
from sqlalchemy.orm import sessionmaker
from sqlalchemy import inspect
import os
import io
import base64
from io import BytesIO
from config import Config
SECRET_KEY = Config.SECRET_KEY
UPLOAD_FOLDER = Config.UPLOAD_FOLDER
from app import db
from app.models.user_models import UploadHistory , CountryProfile , User
from app.utils.data_utils import generate_pnl_report 
from app.utils.us_process_utils  import process_skuwise_us_data , process_us_yearly_skuwise_data, process_us_quarterly_skuwise_data
from app.utils.uk_process_utils import process_skuwise_data , process_quarterly_skuwise_data, process_yearly_skuwise_data 
from app.utils.plotting_utils import (
    get_referral_fees , apply_modifications 
)
from app.utils.currency_utils import (  process_global_yearly_skuwise_data ,
    process_global_quarterly_skuwise_data , 
    process_global_monthly_skuwise_data
)

from dotenv import load_dotenv
from sqlalchemy import MetaData
from sqlalchemy import text



load_dotenv()
db_url = os.getenv('DATABASE_URL')
db_url1= os.getenv('DATABASE_ADMIN_URL')
upload_bp = Blueprint('upload_bp', __name__)

def encode_file_to_base64(file_path):
    with open(file_path, 'rb') as file:
        return base64.b64encode(file.read()).decode('utf-8')




COLUMN_MAPPING = {
    'date/time': 'date_time',
    'settlement id': 'settlement_id',
    'type': 'type',
    'order id': 'order_id',
    'sku': 'sku',
    'description': 'description',
    'quantity': 'quantity',
    'marketplace': 'marketplace',
    'fulfilment': 'fulfilment',
    'fulfillment': 'fulfillment',
    'order city': 'order_city',
    'order state': 'order_state',
    'order postal': 'order_postal',
    'tax collection model': 'tax_collection_model',
    'product sales': 'product_sales',
    'product sales tax': 'product_sales_tax',
    'postage credits': 'postage_credits',
    'shipping credits tax': 'shipping_credits_tax',
    'gift wrap credits': 'gift_wrap_credits',
    'giftwrap credits tax': 'giftwrap_credits_tax',
    'promotional rebates': 'promotional_rebates',
    'promotional rebates tax': 'promotional_rebates_tax',
    'sales tax collected': 'sales_tax_collected',
    'marketplace withheld tax': 'marketplace_facilitator_tax',
    'selling fees': 'selling_fees',
    'fba fees': 'fba_fees',
    'other transaction fees': 'other_transaction_fees',
    'other': 'other',
    'total': 'total',
    'account type': 'account_type',
    'Regulatory Fee': 'regulatory_fee',
    'Tax On Regulatory Fee': 'tax_on_regulatory_fee',
    'Bucket': 'bucket',
    'shipping credits': 'shipping_credits',
    'regulatory fee': 'regulatory_fee',
    'tax on regulatory fee': 'tax_on_regulatory_fee',

}

MONTHS_REVERSE_MAP = {
    1: "january", 2: "february", 3: "march", 4: "april", 5: "may", 6: "june",
    7: "july", 8: "august", 9: "september", 10: "october", 11: "november", 12: "december"
}


MONTHS_MAP = {
    'january': 1, 'february': 2, 'march': 3, 'april': 4,
    'may': 5, 'june': 6, 'july': 7, 'august': 8,
    'september': 9, 'october': 10, 'november': 11, 'december': 12
}

def get_next_month_year(month, year):
    year = int(year)
    month_num = MONTHS_MAP[month]
    if month_num == 12:
        month_num = 1
        month_next = MONTHS_REVERSE_MAP[month_num]
       

        return month_next, year + 1
    month_next = month_num + 1
    month_next1 = MONTHS_REVERSE_MAP[month_next]
    

    return month_next1, year

def table_exists(engine, table_name, schema='public'):
    inspector = inspect(engine)
    return inspector.has_table(table_name, schema=schema)


@upload_bp.route('/upload', methods=['GET', 'POST'])
def upload():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Authorization token is missing or invalid'}), 401

    token = auth_header.split(' ')[1]
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=['HS256'])
        user_id = payload['user_id']
    except jwt.ExpiredSignatureError:
        return jsonify({'error': 'Token has expired'}), 401
    except jwt.InvalidTokenError:
        return jsonify({'error': 'Invalid token'}), 401

    # Retrieve user using SQLAlchemy ORM
    user = User.query.get(user_id)
    if user is None:
        print(f"[DEBUG] User with ID {user_id} not found in database")  # Debug output
        return jsonify({'error': 'User not found'}), 404

    if 'file1' not in request.files or 'file2' not in request.files:
        return jsonify({'success': False, 'message': 'Both files are required'}), 400
    
    file1 = request.files['file1']
    file2 = request.files['file2']
    
    country = request.form['country'].lower()
    month = request.form['month'].lower()
    if month not in ['january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december']:
        return jsonify({'error': 'Invalid month provided'}), 400
    profile = request.form['profile_id']
    year = request.form['year']

    profile = CountryProfile.query.filter(
    db.func.lower(CountryProfile.country) == country,
    CountryProfile.user_id == user_id).first()

    transit_time = int(profile.transit_time)  # Transit time in months
    stock_unit = int(profile.stock_unit)

    table_name = f"user_{user_id}_{country}_{month}{year}_data".lower()
    country_table_name = f"sku_{user_id}_data_table"
    consolidated_table_name = f"user_{user_id}_{country}_merge_data_of_all_months"
    inventory_file_name = f"user_{user_id}_{country}_{month}{year}_inventory_file.xlsx"
    currency_table = f"currency_conversion"
    countris_table_name = f"user_{user_id}_{country}_table"


    # Create SQLAlchemy engine with PostgreSQL
    engine = create_engine(db_url)
    engine1 = create_engine(db_url1)
    meta = MetaData()

    with engine.connect() as connection:
        connection.execute(text(f"DROP TABLE IF EXISTS {table_name}"))
        connection.commit()

    # Define the table - keep the original column case
    user_monthly_data = Table(
        table_name, meta,
        Column('id', Integer, primary_key=True),
        Column('date_time', String),
        Column('settlement_id', String),
        Column('type', String),
        Column('order_id', String),
        Column('sku', String),
        Column('description', String),
        Column('quantity', Integer),
        Column('price_in_gbp', Float),
        Column('cost_of_unit_sold', Float),
        Column('marketplace', String),
        Column('account_type', String),
        Column('fulfilment', String),
        Column('fulfillment', String),
        Column('order_city', String),
        Column('order_state', String),
        Column('order_postal', String),
        Column('tax_collection_model', String),
        Column('regulatory_fee', Float),
        Column('tax_on_regulatory_fee', Float),
        Column('bucket', String),
        Column('product_sales', Float),
        Column('product_sales_tax', Float),
        Column('postage_credits', Float),
        Column('shipping_credits', Float),
        Column('shipping_credits_tax', Float),
        Column('gift_wrap_credits', Float),
        Column('giftwrap_credits_tax', Float),
        Column('promotional_rebates', Float),
        Column('promotional_rebates_tax', Float),
        Column('sales_tax_collected', Float),
        Column('marketplace_withheld_tax', Float),
        Column('marketplace_facilitator_tax', Float),
        Column('selling_fees', Float),
        Column('percentage1', Float),
        Column('fba_fees', Float),
        Column('percentage2', Float),
        Column('other_transaction_fees', Float),
        Column('other', Float),
        Column('total', Float),
        Column('product_name', String),
        Column('currency', String),
        Column('advertising_cost', Float),
        Column('net_reimbursement', Float),
        Column('platform_fees', Float),
        Column('product_group', String),
    )

    user_consolidated_data = Table(
        consolidated_table_name, meta,
        Column('id', Integer, primary_key=True),
        Column('date_time', String),
        Column('settlement_id', String),
        Column('type', String),
        Column('order_id', String),
        Column('sku', String),
        Column('description', String),
        Column('quantity', Integer),
        Column('price_in_gbp', Float),
        Column('cost_of_unit_sold', Float),
        Column('marketplace', String),
        Column('account_type', String),
        Column('fulfilment', String),
        Column('fulfillment', String),
        Column('order_city', String),
        Column('order_state', String),
        Column('order_postal', String),
        Column('tax_collection_model', String),
        Column('regulatory_fee', Float),
        Column('tax_on_regulatory_fee', Float),
        Column('bucket', String),
        Column('product_sales', Float),
        Column('product_sales_tax', Float),
        Column('postage_credits', Float),
        Column('shipping_credits', Float),
        Column('shipping_credits_tax', Float),
        Column('gift_wrap_credits', Float),
        Column('giftwrap_credits_tax', Float),
        Column('promotional_rebates', Float),
        Column('promotional_rebates_tax', Float),
        Column('sales_tax_collected', Float),
        Column('marketplace_withheld_tax', Float),
        Column('marketplace_facilitator_tax', Float),
        Column('selling_fees', Float),
        Column('percentage1', Float),
        Column('fba_fees', Float),
        Column('percentage2', Float),
        Column('other_transaction_fees', Float),
        Column('other', Float),
        Column('total', Float),
        Column('month', String),
        Column('year', String),
        Column('product_name', String),
        Column('currency', String),
        Column('advertising_cost', Float),
        Column('net_reimbursement', Float),
        Column('platform_fees', Float),
        Column('product_group', String),
    )
    meta.create_all(engine)

    with engine.connect() as connection:
        connection.execute(user_monthly_data.delete())  # Delete previous records
        connection.execute(text(f"DELETE FROM {user_consolidated_data} WHERE month = '{month}' AND year = '{year}'"))
        connection.commit()

    # Handle the file upload
    if file1.filename == '' or file2.filename == '':
        return jsonify({'error': 'Both files must be uploaded'}), 400
    
    # Save file1
    file1_path = os.path.join(UPLOAD_FOLDER, f'user_{user_id}_{country}_{month}_{year}_mtd_file.xlsx')
    file1.save(file1_path)
    
    # Save file2 (inventory file)
    if file2.filename.endswith('.csv'):
        # Read CSV file and save as Excel
        df_file2 = pd.read_csv(file2)
        file2_path = os.path.join(UPLOAD_FOLDER, f'user_{user_id}_{country}_{month}{year}_inventory_file.xlsx')
        df_file2.to_excel(file2_path, index=False, engine='openpyxl')  # Save as Excel file
    else:
        # Save the file directly if it's already an Excel file
        file2_path = os.path.join(UPLOAD_FOLDER, secure_filename(inventory_file_name))
        file2.save(file2_path)

    print(f"Saved file1_path: {file1_path}")
    print(f"Saved file2_path: {file2_path}")

    if file1.filename.endswith('.csv'):
        with open(file1_path, 'r', encoding='utf-8-sig') as f:
            lines = f.readlines()
            header_line = None
            # Look for header row
            for i, line in enumerate(lines):
                if any(keyword in line.lower() for keyword in ['date/time', 'date / time', 'date time']):
                    header_line = i
                    break
        if header_line is None:
            return jsonify({'error': 'Could not find header row with "date/time" column'}), 400

        df = pd.read_csv(file1_path, skiprows=header_line, encoding="utf-8-sig", dayfirst=True, skip_blank_lines=True)  # Use utf-8-sig to remove BOM
        df.columns = df.columns.str.strip()
        # Replace problematic characters

    elif file1.filename.endswith(('.xls', '.xlsx')):
        import openpyxl
    
        wb = openpyxl.load_workbook(file1_path, read_only=True)
        ws = wb.active
        header_line = None
        for i, row in enumerate(ws.iter_rows(values_only=True), start=0):
            if row is not None and any(
                cell is not None and any(keyword in str(cell).lower() for keyword in ['date/time', 'date / time', 'date time'])
                for cell in row
            ):
                header_line = i
                break
        if header_line is None:
            return jsonify({'error': 'Could not find header row with "date/time" column in Excel'}), 400

        # Now read excel from the detected header line
        df = pd.read_excel(file1_path, skiprows=header_line)
        df.columns = df.columns.str.strip()
        print(df.head())  #
    else:
        return jsonify({'error': 'Invalid file format. Only .csv and .xlsx files are allowed'}), 400
    
    df.columns = [c.lower() for c in df.columns]  # Lowercase all column names for consistency

    # Function to clean numeric columns
    def clean_numeric_value(val):
        if isinstance(val, str):
            # Remove commas and other non-numeric characters
            if ',' in val:
                val = val.replace(',', '')
            try:
                return float(val)
            except ValueError:
                return None
        return val

    # Apply cleaning function to all numeric columns
    numeric_columns = ['quantity', 'price_in_gbp', 'cost_of_unit_sold', 'product_sales', 
                       'product_sales_tax', 'postage_credits', 'shipping_credits_tax',
                       'gift_wrap_credits', 'giftwrap_credits_tax', 'promotional_rebates',
                       'promotional_rebates_tax', 'sales_tax_collected', 'marketplace_withheld_tax',
                       'marketplace_facilitator_tax', 'selling_fees', 'percentage1', 'fba_fees',
                       'percentage2', 'other_transaction_fees', 'other', 'total','advertising_cost',
                        'net_reimbursement',
                        'platform_fees']
    
    # Make sure all numeric columns exist in the dataframe
    numeric_columns = [col for col in numeric_columns if col in df.columns]
    
    for col in numeric_columns:
        df[col] = df[col].apply(clean_numeric_value)

    df.rename(columns=COLUMN_MAPPING, inplace=True)
    print(df.columns.tolist())
    for col in COLUMN_MAPPING.values():
        if col not in df.columns:
            if col in numeric_columns:
                df[col] = 0
            else:
                df[col] = None  # or use np.nan if preferred
    
    
    if country.upper() == 'UK':
        sku_column = 'sku_uk'
    elif country.upper() == 'US':
        sku_column = 'sku_us'
    else:
        raise ValueError("Unsupported country")

            
    # with engine.connect() as conn:
    #     country_df = pd.read_sql(f"SELECT {sku_column} AS sku, price,currency, product_name FROM {country_table_name}", conn)

    # df = df.merge(country_df, on='sku', how='left')

    

    MONTH_NUM = {
        "january": 1, "february": 2, "march": 3, "april": 4,
        "may": 5, "june": 6, "july": 7, "august": 8,
        "september": 9, "october": 10, "november": 11, "december": 12
    }

    target_month_num = MONTH_NUM.get(month.lower())
    if not target_month_num:
        raise ValueError(f"Invalid month: {month}. Expected january..december")

    target_year = int(year)  # form year string -> int

    month_case_sql = """
    CASE lower(month)
        WHEN 'january' THEN 1
        WHEN 'february' THEN 2
        WHEN 'march' THEN 3
        WHEN 'april' THEN 4
        WHEN 'may' THEN 5
        WHEN 'june' THEN 6
        WHEN 'july' THEN 7
        WHEN 'august' THEN 8
        WHEN 'september' THEN 9
        WHEN 'october' THEN 10
        WHEN 'november' THEN 11
        WHEN 'december' THEN 12
        ELSE NULL
    END
    """

    # year is string in DB, so cast it for comparisons
    price_asof_query = text(f"""
        SELECT sku, price, currency, product_name
        FROM (
            SELECT
                {sku_column} AS sku,
                price,
                currency,
                product_name,
                CAST(year AS INTEGER) AS year_int,
                {month_case_sql} AS month_num,
                ROW_NUMBER() OVER (
                    PARTITION BY {sku_column}
                    ORDER BY CAST(year AS INTEGER) DESC, {month_case_sql} DESC
                ) AS rn
            FROM {country_table_name}
            WHERE
                year IS NOT NULL
                AND trim(year) <> ''
                AND {month_case_sql} IS NOT NULL
                AND (
                    CAST(year AS INTEGER) < :target_year
                    OR (CAST(year AS INTEGER) = :target_year AND {month_case_sql} <= :target_month_num)
                )
        ) x
        WHERE x.rn = 1
    """)

    with engine.connect() as conn:
        country_df = pd.read_sql(
            price_asof_query,
            conn,
            params={"target_year": target_year, "target_month_num": target_month_num}
        )

    df = df.merge(country_df, on="sku", how="left")


    

    with engine.connect() as conn:
        countries_df = pd.read_sql(f"SELECT sku, product_group FROM {countris_table_name}", conn)

    df = df.merge(countries_df, on='sku', how='left')

    with engine1.connect() as conn:
        currency_query = text("""
            SELECT conversion_rate
            FROM currency_conversion 
            WHERE lower(user_currency) = :currency 
            AND lower(country) = :country 
            AND lower(month) = :month 
            AND year = :year
            LIMIT 1
        """)

        result = conn.execute(
            currency_query,
            {
                "currency": country_df['currency'].dropna().iloc[0].lower(),  # Get any available currency from df
                "country": country.lower(),
                "month": month.lower(),
                "year": year
            }
        ).fetchone()

# Step 3: Apply conversion rate to calculate price_in_gbp
    conversion_rate = result[0] if result else None

    if conversion_rate:
        df['price_in_gbp'] = df['price'] * conversion_rate
    else:
        df['price_in_gbp'] = None  # Or handle fallback logic if conversion rate not found

    
    # Calculate cost_of_unit_sold where both values are not null
    df['cost_of_unit_sold'] = df.apply(lambda row: row['price_in_gbp'] * row['quantity'] 
                                      if pd.notnull(row['price_in_gbp']) and pd.notnull(row['quantity']) 
                                      else None, axis=1)

    df.drop(columns=['price'], inplace=True)

    # Clean the 'total' column specifically since it appears in the error message
    if 'total' in df.columns:
        df['total'] = df['total'].apply(clean_numeric_value)

    # Handle the case where any numeric column might still have commas or other formatting issues
    for col in [c for c in df.columns if c in numeric_columns]:
        if col in df.columns:
            # Convert to float with NaN for any unconvertible values
            df[col] = pd.to_numeric(df[col], errors='coerce')

    for col in df.columns:
                if df[col].dtype == 'object':  # likely a string column
                    if df[col].str.contains(',').any():
                        print(f"Column {col} contains comma-formatted numbers!")
    for col in df.columns:
                if df[col].dtype == 'object':  # likely to have commas or bad data
                    # Remove commas and convert to numeric
                    df[col] = df[col].str.replace(',', '', regex=True)

            # Now convert all possible numeric columns to float
    df = df.apply(lambda x: pd.to_numeric(x, errors='ignore') if x.dtype == 'object' else x)

    df.to_sql(table_name, con=engine, if_exists='append', index=False)
    
    df['month'] = month  # Assigning month from form data
    df['year'] = year 

    for col in df.columns:
                if df[col].dtype == 'object':  # likely a string column
                    if df[col].str.contains(',').any():
                        print(f"Column {col} contains comma-formatted numbers!")
    for col in df.columns:
                if df[col].dtype == 'object':  # likely to have commas or bad data
                    # Remove commas and convert to numeric
                    df[col] = df[col].str.replace(',', '', regex=True)

            # Now convert all possible numeric columns to float
    df = df.apply(lambda x: pd.to_numeric(x, errors='ignore') if x.dtype == 'object' else x)

    
    

    df.to_sql(consolidated_table_name, con=engine, if_exists='append', index=False)


    # ✅ Step 1: Currency Rates Dictionary
    if country.lower() == 'uk':
        currency1 = 'gbp'
    elif country.lower() == 'us':
        currency1 = 'usd'
    elif country.lower() == 'canada':
        currency1 = 'cad'
    else:
        currency1 = 'usd'  # fallback/default if unknown

    # Step 2: Fetch conversion rate from currency_conversion table
    with engine1.connect() as conn:
        currency_query = text("""
            SELECT conversion_rate
            FROM currency_conversion 
            WHERE lower(user_currency) = :currency1
            AND lower(country) = 'us'
            AND lower(month) = :month 
            AND year = :year
            LIMIT 1
        """)
        
        result = conn.execute(currency_query, {
            "currency1": currency1,
            "country": "us",
            "month": month.lower(),
            "year": year
        }).fetchone()

    # Step 3: Use the conversion rate
    currency_rate  = result[0] if result else None

    # Step 4: Create a USD-converted copy if conversion rate exists
    df_usd = df.copy()

    # ✅ Step 3: Convert monetary columns to USD if country found
    if currency_rate :
        monetary_columns = [
            'product_sales', 'product_sales_tax', 'postage_credits', 
            'shipping_credits_tax', 'gift_wrap_credits', 'giftwrap_credits_tax', 
            'promotional_rebates', 'promotional_rebates_tax', 'sales_tax_collected',
            'marketplace_facilitator_tax', 'selling_fees', 'fba_fees', 
            'other_transaction_fees', 'other', 'total', 'price_in_gbp', 
            'cost_of_unit_sold'
        ]

        for col in monetary_columns:
            if col in df_usd.columns:
                df_usd[col] = pd.to_numeric(df_usd[col], errors='coerce') * currency_rate 
    else:
        print("⚠️ No conversion rate found for:", currency1, country, month, year)

    # ✅ Step 4: Insert the USD converted data into global table
    global_table_name = f"user_{user_id}_total_country_global_data"

    user_global_table = Table(
        global_table_name, meta,
        Column('id', Integer, primary_key=True),  # optional, just for reflection
        Column('date_time', String),  # Changed from DateTime to String
        Column('settlement_id', String),
        Column('type', String),
        Column('order_id', String),
        Column('sku', String),
        Column('description', String),
        Column('quantity', Integer),
        Column('price_in_gbp', Float),
        Column('cost_of_unit_sold', Float),
        Column('marketplace', String),
        Column('fulfilment', String),
        Column('fulfillment', String),
        Column('order_city', String),
        Column('order_state', String),
        Column('order_postal', String),
        Column('tax_collection_model', String),
        Column('product_sales', Float),
        Column('product_sales_tax', Float),
        Column('postage_credits', Float),
        Column('shipping_credits', Float),
        Column('shipping_credits_tax', Float),
        Column('gift_wrap_credits', Float),
        Column('giftwrap_credits_tax', Float),
        Column('promotional_rebates', Float),
        Column('promotional_rebates_tax', Float),
        Column('sales_tax_collected', Float),
        Column('marketplace_withheld_tax', Float),
        Column('marketplace_facilitator_tax', Float),
        Column('selling_fees', Float),
        Column('percentage1', Float),
        Column('fba_fees', Float),
        Column('percentage2', Float),
        Column('other_transaction_fees', Float),
        Column('other', Float),
        Column('total', Float),
        Column('month', String),  # Added column
        Column('year', String),
        Column('product_name', String),
        Column('country', String),
        
    )
    meta.create_all(engine)

    with engine.connect() as connection:
        # Delete previous records from user_total_country_global_data
        connection.execute(
            text(f"DELETE FROM {global_table_name} WHERE month = :month AND year = :year AND country = :country"),
            {"month": month, "year": year, "country": country}
        )
        connection.commit()

    # Add month, year, and country columns to the dataframe
    df_usd['month'] = month
    df_usd['year'] = year
    df_usd['country'] = country

    # Ensure all data types are correct before insertion
    # Convert any problematic columns to appropriate types
    for col in df_usd.columns:
        if df_usd[col].dtype == 'object' and col not in ['date_time', 'settlement_id', 'type', 'order_id', 'sku', 
                                                        'description', 'marketplace', 'fulfilment', 'order_city', 
                                                        'order_state', 'order_postal', 'tax_collection_model',
                                                        'month', 'year', 'country', 'product_name']:
            # Try to convert string columns that should be numeric to float
            df_usd[col] = pd.to_numeric(df_usd[col], errors='coerce')

    # Handle NaN values explicitly
    df_usd = df_usd.fillna({col: 0.0 for col in df_usd.select_dtypes(include=['float64']).columns})
    df_usd = df_usd.fillna({col: 0 for col in df_usd.select_dtypes(include=['int64']).columns})
    df_usd = df_usd.fillna('')  # Fill string columns with empty string

    try:
        # Convert DataFrame to dict of records for proper SQL insertion
        records = df_usd.to_dict(orient='records')
        
        with engine.begin() as connection:
            for chunk in [records[i:i+1000] for i in range(0, len(records), 1000)]:
                connection.execute(user_global_table.insert(), chunk)
        
        print(f"Successfully inserted {len(df_usd)} records into {global_table_name}")
    except Exception as e:
        print(f"Error during insertion: {str(e)}")
        # Print more debugging information
        print(f"DataFrame info: {df_usd.info()}")
        print(f"First row sample: {df_usd.iloc[0].to_dict() if len(df_usd) > 0 else 'Empty DataFrame'}")

    # Update required columns to match the actual column names in the database
    REQUIRED_COLUMNS = [
            "order_id", "sku", "description", "product_sales", "product_sales_tax", 
            "postage_credits", "shipping_credits_tax", "gift_wrap_credits", "giftwrap_credits_tax", 
            "promotional_rebates", "promotional_rebates_tax", "marketplace_facilitator_tax", 
            "selling_fees", "fba_fees", "other_transaction_fees", "errorstatus", "answer", 
            "difference", "fbaerrorstatus", "fbaanswer"  # Use original capitalization
    ]
        
    # Convert both files to base64
    def encode_file_to_base64(file_path):
        with open(file_path, "rb") as file:
            return base64.b64encode(file.read()).decode()
             
    def replace_nan_with_null(data):
        if isinstance(data, dict):
            return {key: replace_nan_with_null(value) for key, value in data.items()}
        elif isinstance(data, list):
            return [replace_nan_with_null(item) for item in data]
        elif isinstance(data, float) and (data != data):  # NaN check
            return None
        return data
        
    if file1 and file1.filename != '':
        try:
            if file1.filename.endswith('.csv'):
                with open(file1_path, 'r', encoding='utf-8-sig') as f:
                    lines = f.readlines()
                    header_line = None
                    # Look for header row
                    for i, line in enumerate(lines):
                        if any(keyword in line.lower() for keyword in ['date/time', 'date / time', 'date time']):
                            header_line = i
                            break
                if header_line is None:
                    return jsonify({'error': 'Could not find header row with "date/time" column'}), 400

                df = pd.read_csv(file1_path, skiprows=header_line, encoding="utf-8-sig", dayfirst=True, skip_blank_lines=True)
                df.columns = df.columns.str.strip()

            elif file1.filename.endswith(('.xls', '.xlsx')):
                import openpyxl
    
                wb = openpyxl.load_workbook(file1_path, read_only=True)
                ws = wb.active
                header_line = None
                for i, row in enumerate(ws.iter_rows(values_only=True), start=0):
                    if row is not None and any(
                        cell is not None and any(keyword in str(cell).lower() for keyword in ['date/time', 'date / time', 'date time'])
                        for cell in row
                    ):
                        header_line = i
                        break
                if header_line is None:
                    return jsonify({'error': 'Could not find header row with "date/time" column in Excel'}), 400

                # Now read excel from the detected header line
                df = pd.read_excel(file1_path, skiprows=header_line)
                df.columns = df.columns.str.strip()
                print(df.head())
            else:
                return jsonify({'error': 'Invalid file format. Only .csv and .xlsx files are allowed'}), 400

            df.columns = [c.lower() for c in df.columns]  # Lowercase all column names for consistency
            
            # Clean numeric columns again for this new dataframe
            for col in [c for c in df.columns if c in numeric_columns]:
                df[col] = df[col].apply(clean_numeric_value)
                df[col] = pd.to_numeric(df[col], errors='coerce')
                
            df.rename(columns=COLUMN_MAPPING, inplace=True)
            for col in COLUMN_MAPPING.values():
                if col not in df.columns:
                    if col in numeric_columns:
                        df[col] = 0
                    else:
                        df[col] = None  # or use np.nan if preferred
            
            if country.upper() == 'UK':
                sku_column = 'sku_uk'
            elif country.upper() == 'US':
                sku_column = 'sku_us'
            else:
                raise ValueError("Unsupported country")

            
            # with engine.connect() as conn:
            #     country_df = pd.read_sql(f"SELECT {sku_column} AS sku, price,currency, product_name FROM {country_table_name}", conn)

            # df = df.merge(country_df, on='sku', how='left')

            
            MONTH_NUM = {
                "january": 1, "february": 2, "march": 3, "april": 4,
                "may": 5, "june": 6, "july": 7, "august": 8,
                "september": 9, "october": 10, "november": 11, "december": 12
            }

            target_month_num = MONTH_NUM.get(month.lower())
            if not target_month_num:
                raise ValueError(f"Invalid month: {month}. Expected january..december")

            target_year = int(year)  # form year string -> int

            month_case_sql = """
            CASE lower(month)
                WHEN 'january' THEN 1
                WHEN 'february' THEN 2
                WHEN 'march' THEN 3
                WHEN 'april' THEN 4
                WHEN 'may' THEN 5
                WHEN 'june' THEN 6
                WHEN 'july' THEN 7
                WHEN 'august' THEN 8
                WHEN 'september' THEN 9
                WHEN 'october' THEN 10
                WHEN 'november' THEN 11
                WHEN 'december' THEN 12
                ELSE NULL
            END
            """

            # year is string in DB, so cast it for comparisons
            price_asof_query = text(f"""
                SELECT sku, price, currency, product_name
                FROM (
                    SELECT
                        {sku_column} AS sku,
                        price,
                        currency,
                        product_name,
                        CAST(year AS INTEGER) AS year_int,
                        {month_case_sql} AS month_num,
                        ROW_NUMBER() OVER (
                            PARTITION BY {sku_column}
                            ORDER BY CAST(year AS INTEGER) DESC, {month_case_sql} DESC
                        ) AS rn
                    FROM {country_table_name}
                    WHERE
                        year IS NOT NULL
                        AND trim(year) <> ''
                        AND {month_case_sql} IS NOT NULL
                        AND (
                            CAST(year AS INTEGER) < :target_year
                            OR (CAST(year AS INTEGER) = :target_year AND {month_case_sql} <= :target_month_num)
                        )
                ) x
                WHERE x.rn = 1
            """)

            with engine.connect() as conn:
                country_df = pd.read_sql(
                    price_asof_query,
                    conn,
                    params={"target_year": target_year, "target_month_num": target_month_num}
                )

            df = df.merge(country_df, on="sku", how="left")


            with engine.connect() as conn:
                countries_df = pd.read_sql(f"SELECT sku, product_group FROM {countris_table_name}", conn)

            df = df.merge(countries_df, on='sku', how='left')

            with engine1.connect() as conn:
                currency_query = text("""
                    SELECT conversion_rate
                    FROM currency_conversion 
                    WHERE lower(user_currency) = :currency 
                    AND lower(country) = :country 
                    AND lower(month) = :month 
                    AND year = :year
                    LIMIT 1
            """)

                result = conn.execute(
                    currency_query,
                    {
                        "currency": country_df['currency'].dropna().iloc[0].lower(),  # Get any available currency from df
                        "country": country.lower(),
                        "month": month.lower(),
                        "year": year
                    }
                ).fetchone()

        # Step 3: Apply conversion rate to calculate price_in_gbp
            conversion_rate = result[0] if result else None

            if conversion_rate:
                df['price_in_gbp'] = df['price'] * conversion_rate
            else:
                df['price_in_gbp'] = None  # Or handle fallback logic if conversion rate not found

            # Calculate cost_of_unit_sold where both values are not null
            df['cost_of_unit_sold'] = df.apply(lambda row: row['price_in_gbp'] * row['quantity'] 
                                          if pd.notnull(row['price_in_gbp']) and pd.notnull(row['quantity']) 
                                          else None, axis=1)

            df.drop(columns=['price'], inplace=True)

            # Convert any remaining string numeric values to float
            for col in numeric_columns:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')

            for col in df.columns:
                if df[col].dtype == 'object':  # likely a string column
                    if df[col].str.contains(',').any():
                        print(f"Column {col} contains comma-formatted numbers!")
            for col in df.columns:
                        if df[col].dtype == 'object':  # likely to have commas or bad data
                            # Remove commas and convert to numeric
                            df[col] = df[col].str.replace(',', '', regex=True)

                    # Now convert all possible numeric columns to float
            df = df.apply(lambda x: pd.to_numeric(x, errors='ignore') if x.dtype == 'object' else x)

            # df.to_sql(table_name, con=engine, if_exists='append', index=False)
            df.to_sql(table_name, con=engine, if_exists='append', index=False)

            file1_base64 = encode_file_to_base64(file1_path)
            file2_base64 = encode_file_to_base64(file2_path)

            sku_list = df['sku'].dropna().tolist()  # Filter out NaN values

            referral_fees = get_referral_fees(user_id, country, sku_list)

            if referral_fees is not None:
                df_modified = apply_modifications(df, country)
                
                # Ensure numeric columns are properly converted before saving to database
                for col in numeric_columns:
                    if col in df_modified.columns:
                        df_modified[col] = pd.to_numeric(df_modified[col], errors='coerce')
                
                df_modified.to_sql(table_name, con=engine, if_exists='replace', index=False, method='multi')

                excel_output = io.BytesIO()
                df_modified.to_excel(excel_output, index=False)
                excel_output.seek(0)

                pnl_report = generate_pnl_report(year, month)

                table_name = f"user_{user_id}_{country}_{month}{year}_data"
    
                with engine.connect() as conn:
                    query = f"""
                    SELECT * FROM {table_name} 
                    WHERE ErrorStatus IN ('cases to be inquired', 'NoReferralFee')
                    AND sku <> '0'
                    AND sku IS NOT NULL
                    AND TRIM(sku) <> ''
                    """
                    print(f"Executing query on table: {table_name}")  # Debugging
                    error_df = pd.read_sql(query, conn)
                
                # Filter error_df to keep only the required columns
                error_df = error_df[[col for col in REQUIRED_COLUMNS if col in error_df.columns]]

                # Save the error file if there are errors
                error_file_path = None
                error_file_base64 = None
                if not error_df.empty:
                    error_filename = f"error_file_{user_id}{country}{month}_{year}.xlsx"
                    error_file_path = os.path.join(UPLOAD_FOLDER, error_filename)                    
                    error_df.to_excel(error_file_path, index=False)
                    
                    # Encode error file to base64
                    with open(error_file_path, "rb") as error_file:
                        error_file_base64 = base64.b64encode(error_file.read()).decode()

                quarter_mapping = {
                    'january': 'Q1', 'february': 'Q1', 'march': 'Q1',
                    'april': 'Q2', 'may': 'Q2', 'june': 'Q2',
                    'july': 'Q3', 'august': 'Q3', 'september': 'Q3',
                    'october': 'Q4', 'november': 'Q4', 'december': 'Q4'
                }

                if month.lower() not in quarter_mapping:
                    print(f"Error: Invalid month {month}")  # Debugging
                    return jsonify({'error': 'Invalid month provided'}), 400

                quarter = quarter_mapping[month.lower()]
                quarterly_table = f"quater{quarter[-1]}{country}{year}_table"
                print(f"Determined quarterly table: {quarterly_table}")  # Debugging
                
                # Make sure these function calls use engine properly
                # qtd_pie_chart = process_quarterly_skuwise_data(user_id, country, month, year, quarter, db_url)
                # ytd_pie_chart = process_yearly_skuwise_data(user_id, country, year)
                print("Quarterly table updated successfully")  # Debugging
                
                
                   

                # Generate sales pie chart
                if country.lower() == 'uk':
                    total_cous, total_amazon_fee, cm2_profit, rembursement_fee, platform_fee, total_expense, total_profit, total_fba_fees, advertising_total, taxncredit, reimbursement_vs_sales, cm2_margins, acos, rembursment_vs_cm2_margins, total_sales, unit_sold  = process_skuwise_data(user_id, country, month, year)
                    ytd_pie_chart = process_yearly_skuwise_data(user_id, country, year)
                    qtd_pie_chart = process_quarterly_skuwise_data(user_id, country, month, year, quarter, db_url)
                    # sales_pie_chart, platform_fee, rembursement_fee = create_sales_pie_chart(df_modified)
                    # expense_pie_chart,   otherwplatform = create_expense_pie_chart(df_modified, country, month, year)
                    
                elif country.lower() == 'us':
                    platform_fee, rembursement_fee, total_cous, total_amazon_fee,  total_profit, total_expense, total_fba_fees, cm2_profit, cm2_margins, acos, rembursment_vs_cm2_margins, advertising_total, reimbursement_vs_sales, unit_sold, total_sales, otherwplatform, taxncredit = process_skuwise_us_data(user_id, country, month, year)
                    ytd_pie_chart = process_us_yearly_skuwise_data(user_id, country, year)
                    qtd_pie_chart = process_us_quarterly_skuwise_data(user_id, country, month, year, quarter, db_url)
                

                process_global_monthly_skuwise_data(user_id, country, year, month)
                process_global_quarterly_skuwise_data(user_id, country, month, year, quarter , db_url)
                process_global_yearly_skuwise_data(user_id, country, year)


                
                
                # Check if entry already exists using string month
                existing_entry = UploadHistory.query.filter_by(user_id=user_id, country=country, month=month, year=year).first()
                if existing_entry:
                    db.session.delete(existing_entry)
                    db.session.commit()

                new_upload = UploadHistory(
                    user_id=user_id,
                    year=year,
                    month=month,
                    country=country,
                    file_name=secure_filename(file1.filename), 
                    sales_chart_img=None,
                    expense_chart_img=None,
                    total_sales=float(total_sales),
                    total_profit=float(total_profit),
                    otherwplatform=float(platform_fee),
                    taxncredit = float(taxncredit) if taxncredit is not None else 0.0,
                    total_expense=float(total_expense),
                    qtd_pie_chart=qtd_pie_chart,
                    ytd_pie_chart=ytd_pie_chart,
                    total_cous=float(total_cous),
                    total_amazon_fee=float(total_amazon_fee),
                    total_fba_fees=float(total_fba_fees),
                    platform_fee=float(platform_fee),  # Include platform_fee here
                    rembursement_fee=float(rembursement_fee),
                    cm2_profit= float(cm2_profit),
                    cm2_margins= float(cm2_margins),
                    acos= float(acos),
                    rembursment_vs_cm2_margins= float(rembursment_vs_cm2_margins),
                    advertising_total= float(advertising_total),
                    reimbursement_vs_sales= float(reimbursement_vs_sales),
                    unit_sold = int(unit_sold)
                )
                db.session.add(new_upload)
                db.session.commit()       

                next_month, next_year = get_next_month_year(month, year)
                next_year = str(next_year)
                next_table = f"skuwisemonthly_{user_id}_{country}_{next_month}{next_year}"
                print(f"Checking table: {next_table}")
                print("run")

                if table_exists(engine, next_table):
                    print("Table exists! Running next month logic...")
    # Run for next month too
                    if country.lower() == 'uk':
                        total_cous, total_amazon_fee, cm2_profit, rembursement_fee, platform_fee, total_expense, total_profit, total_fba_fees, advertising_total, taxncredit, reimbursement_vs_sales, cm2_margins, acos, rembursment_vs_cm2_margins, total_sales, unit_sold = process_skuwise_data(user_id, country, next_month, next_year)
                    elif country.lower() == 'us':
                        platform_fee, rembursement_fee, total_cous, total_amazon_fee,  total_profit, total_expense, total_fba_fees, cm2_profit, cm2_margins, acos, rembursment_vs_cm2_margins, advertising_total, reimbursement_vs_sales, unit_sold, total_sales, otherwplatform,taxncredit = process_skuwise_us_data(user_id, country, next_month, next_year)
                        print("success run")

                else:
                    print("Table does NOT exist for next month.")

                      
            
                response_data = {
                    'success': True,
                    # 'sales_chart_img': sales_pie_chart,
                    # 'expense_chart_img': expense_pie_chart,
                    'total_sales': total_sales,
                    'total_profit': total_profit,
                    'otherwplatform': platform_fee,
                    'taxncredit': taxncredit,
                    'total_expense': total_expense,
                    'total_fba_fees': total_fba_fees,  
                    'excel_file': base64.b64encode(excel_output.getvalue()).decode(),
                    'file1': file1_base64,  # Original file 1 in base64
                    'file2': file2_base64,  # Original file 2 in base64
                    'error_file': error_file_base64 if 'error_file_base64' in locals() and error_file_base64 else None,
                    'platform_fee': platform_fee,
                }
                # Clean up NaN values before returning the response
                response_data_cleaned = replace_nan_with_null(response_data)
                return jsonify(response_data_cleaned)
            else:
                return jsonify({'success': False, 'message': 'Referral fee not found for the SKUs in the uploaded file.'}), 400
        except Exception as e:
            print("Error:", e)
            return jsonify({'success': False, 'message': 'An error occurred while processing your request.'}), 500

    else:
        return jsonify({'success': False, 'message': 'No file selected. Please select a file.'}), 400





# ---------- Helpers ----------

def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Trim, lowercase, replace spaces/dashes with underscores."""
    df.columns = (
        df.columns.astype(str)
        .str.strip()
        .str.lower()
        .str.replace(r'[\s\-]+', '_', regex=True)
    )
    return df

def _promote_first_row_to_header_if_needed(df: pd.DataFrame) -> pd.DataFrame:
    """Detect sheets where pandas used 'Unnamed' headers and real headers are on row 0."""
    total = len(df.columns)
    if total == 0:
        return df
    unnamed_count = sum(str(c).lower().startswith('unnamed') for c in df.columns)
    if unnamed_count / total >= 0.5 and len(df) > 0:
        first_row = df.iloc[0]
        candidate_headers = [str(x).strip() if x is not None else "" for x in first_row.tolist()]
        known = {
            's. no.', 's_no', 'product name', 'product_name',
            'product barcode', 'product_barcode', 'asin', 'sku',
            'sku_uk', 'sku_us', 'sku_canada', 'landing cost',
            'landing_cost', 'your price', 'your_price', 'currency',
            'amazon-store', 'amazon_store', 'marketplace'
        }
        matches = sum(1 for x in candidate_headers if x.lower() in known)
        if matches >= 2:
            df = df.copy()
            df.columns = candidate_headers
            df = df.iloc[1:].reset_index(drop=True)
    return df

def _pick(row: dict, keys):
    for k in keys:
        if k in row and row[k] not in (None, '', '—', '--'):
            return row[k]
    return None

def _marketplace_to_country(store):
    if not store:
        return None
    s = str(store).strip().upper()
    if s in {'GB', 'UK', 'GBR', 'UNITED KINGDOM'}:
        return 'UK'
    if s in {'US', 'USA', 'UNITED STATES'}:
        return 'US'
    return None


@upload_bp.route('/ConfirmationFeepreview', methods=['GET'])
def ConfirmationFeepreview():
    return jsonify({'message': 'ConfirmationFeepreview successful!'}), 200


@upload_bp.route('/multiCountry', methods=['POST'])
def multiCountry():
    # ---------- Auth ----------
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Authorization token is missing or invalid'}), 401

    token = auth_header.split(' ')[1]
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=['HS256'])
        user_id = payload['user_id']
    except jwt.ExpiredSignatureError:
        return jsonify({'error': 'Token has expired'}), 401
    except jwt.InvalidTokenError:
        return jsonify({'error': 'Invalid token'}), 401

    # ---------- File ----------
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'No file provided'}), 400
    if not (file.filename.endswith('.csv') or file.filename.endswith('.xlsx')):
        return jsonify({'error': 'Invalid file type. Only CSV or XLSX files are allowed.'}), 400

    # ---------- DB ----------
    db_url = os.getenv('DATABASE_URL', 'postgresql://postgres:password@localhost:5432/phormula')
    engine = create_engine(db_url)
    inspector = inspect(engine)
    metadata = MetaData()  # SQLAlchemy 2.x: no bind here
    table_name = f"sku_{user_id}_data_table"

    # Relaxed schema (only user_id is NOT NULL)
    user_specific_table = Table(
        table_name, metadata,
        Column('id', Integer, primary_key=True),
        Column('user_id', Integer, nullable=False),
        Column('s_no', Integer, nullable=True),
        Column('product_name', String(255), nullable=True),
        Column('product_barcode', String(255), nullable=True),
        Column('sku_uk', String(255), nullable=True),
        Column('sku_us', String(255), nullable=True),
        Column('asin', String(255), nullable=True),
        Column('price', Float, nullable=True),
        Column('currency', String(255), nullable=True),
        Column('month', String(20), nullable=True),
        Column('year', String(20), nullable=True),

    )

    session = None
    try:
        # ---------- Read into DataFrame ----------
        if file.filename.endswith('.csv'):
            try:
                df = pd.read_csv(BytesIO(file.read()))
            except UnicodeDecodeError:
                file.stream.seek(0)
                df = pd.read_csv(BytesIO(file.read()), encoding='latin-1')
        else:
            df = pd.read_excel(BytesIO(file.read()))

        # Fix “Unnamed” header case, then normalize
        df = _promote_first_row_to_header_if_needed(df)
        df = _normalize_columns(df)
        df = df.where(pd.notnull(df), None)

        # (Re)create fresh table per upload
        if inspector.has_table(table_name):
            Table(table_name, MetaData(), autoload_with=engine).drop(engine, checkfirst=True)
        metadata.create_all(engine)  # SQLAlchemy 2.x: pass engine here

        Session = sessionmaker(bind=engine)
        session = Session()

        inserts = []
        for _, r in df.iterrows():
            row = r.to_dict()

            s_no = _pick(row, ['s_no', 's_no.', 's._no.', 'no', '#'])
            try:
                s_no = int(s_no) if s_no not in (None, '') else None
            except Exception:
                s_no = None

            product_name = _pick(row, ['product_name', 'product-name', 'title', 'item_name'])
            product_barcode = _pick(row, ['product_barcode', 'barcode', 'ean', 'upc'])
            asin = _pick(row, ['asin'])

            amazon_store = _pick(row, ['amazon_store', 'amazon-store', 'marketplace'])
            country = _marketplace_to_country(amazon_store)

            raw_sku = _pick(row, ['sku', 'seller_sku', 'merchant_sku', 'sku_uk', 'sku_us'])
            sku_uk = _pick(row, ['sku_uk'])
            sku_us = _pick(row, ['sku_us'])
            if not sku_uk and not sku_us and raw_sku:
                if country == 'UK':
                    sku_uk = str(raw_sku).strip()
                elif country == 'US':
                    sku_us = str(raw_sku).strip()

            price_value = _pick(row, ['landing_cost', 'your_price', 'sales_price', 'price'])
            try:
                price_value = float(price_value) if price_value not in (None, '') else None
            except Exception:
                price_value = None

            currency = _pick(row, ['currency'])

            month = _pick(row, ['month', 'Month', 'mon', 'mm'])
            year  = _pick(row, ['year', 'Year', 'yyyy', 'yy'])

            # int conversion (safe)
            try:
                month = str(month).strip().lower() if month not in (None, '') else None
            except Exception:
                month = None

            try:
                year = str(year) if year not in (None, '') else None
            except Exception:
                year = None


            # Skip fully empty lines
            if not any([s_no, product_name, product_barcode, asin, sku_uk, sku_us, price_value, currency]):
                continue

            def _s(x):
                if x is None:
                    return None
                x = str(x).strip()
                return x if x else None

            inserts.append({
                'user_id': user_id,
                's_no': s_no,
                'product_name': _s(product_name),
                'product_barcode': _s(product_barcode),
                'sku_uk': _s(sku_uk),
                'sku_us': _s(sku_us),
                'asin': _s(asin),
                'price': price_value,
                'currency': _s(currency),
                'month': month,
                'year': year,
            })

        if inserts:
            session.execute(user_specific_table.insert(), inserts)
            session.commit()
            msg = 'File uploaded and data saved successfully'
        else:
            msg = 'File processed, but no valid rows found to insert.'

        session.close()
        engine.dispose()
        return jsonify({'message': msg}), 200

    except Exception as e:
        try:
            if session is not None:
                session.rollback()
                session.close()
        except Exception:
            pass
        try:
            engine.dispose()
        except Exception:
            pass
        print(f"Error processing file: {str(e)}")
        return jsonify({'error': f'Error processing file: {str(e)}'}), 500


@upload_bp.route('/file-upload-status', methods=['GET'])
def check_file_upload_status():
    # ---------- Auth ----------
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Authorization token is missing or invalid'}), 401

    token = auth_header.split(' ')[1]
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=['HS256'])
        user_id = payload['user_id']
    except jwt.ExpiredSignatureError:
        return jsonify({'error': 'Token has expired'}), 401
    except jwt.InvalidTokenError:
        return jsonify({'error': 'Invalid token'}), 401

    # ---------- DB ----------
    db_url = os.getenv('DATABASE_URL', 'postgresql://postgres:password@localhost:5432/phormula')
    engine = create_engine(db_url)
    inspector = inspect(engine)
    table_name = f"sku_{user_id}_data_table"

    try:
        if not inspector.has_table(table_name):
            engine.dispose()
            return jsonify({'file_uploaded': False}), 200

        with engine.connect() as conn:
            result = conn.execute(
                text(f"""
                    SELECT COUNT(*) FROM {table_name}
                    WHERE user_id = :user_id
                      AND (sku_us IS NOT NULL OR sku_uk IS NOT NULL)
                """),
                {"user_id": user_id}
            ).scalar() or 0

        engine.dispose()
        return jsonify({'file_uploaded': result > 0}), 200

    except Exception as e:
        print(f"Error checking file upload status: {str(e)}")
        try:
            engine.dispose()
        except Exception:
            pass
        return jsonify({'error': 'Server error'}), 500




from flask import request, jsonify
import jwt

@upload_bp.route('/upload_history', methods=['GET'])
def upload_history():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Authorization token is missing or invalid'}), 401

    token = auth_header.split(' ')[1]
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=['HS256'])
        user_id = payload['user_id']
    except jwt.ExpiredSignatureError:
        return jsonify({'error': 'Token has expired'}), 401
    except jwt.InvalidTokenError:
        return jsonify({'error': 'Invalid token'}), 401

    # ✅ read optional params
    # If you also pass country from FE, support it too (safe default = "")
    country_param = (request.args.get('country') or "").strip().lower()
    home_currency = (request.args.get('homeCurrency') or "").strip().lower()

    uploads = UploadHistory.query.filter_by(user_id=user_id).all()

    # Month number to name mapping
    month_names = {
        1: 'january', 2: 'february', 3: 'march', 4: 'april',
        5: 'may', 6: 'june', 7: 'july', 8: 'august',
        9: 'september', 10: 'october', 11: 'november', 12: 'december'
    }

    response = []

    for upload in uploads:
        upload_country = (upload.country or "").strip().lower()

        # ✅ IMPORTANT FIX:
        # When requesting GLOBAL history:
        # - if homeCurrency is provided => only return global_<currency>
        # - else => only return base global
        if country_param == "global":
            if home_currency:
                if upload_country != f"global_{home_currency}":
                    continue
            else:
                if upload_country != "global":
                    continue

        # ✅ Optional: if FE passes specific country, filter by it
        # (This avoids sending huge payloads)
        elif country_param:
            if upload_country != country_param:
                continue

        # Convert numeric month to month name for display
        month_name = month_names.get(upload.month, str(upload.month))

        table_name = f"user_{upload_country}_{month_name}{upload.year}_data"

        response.append({
            'month': month_name,
            'month_num': upload.month,          # ✅ keep numeric month here
            'year': upload.year,
            'country': upload_country,
            'file_name': table_name,

            'total_sales': upload.total_sales,
            'total_profit': upload.total_profit,
            'total_expense': upload.total_expense,
            'total_fba_fees': upload.total_fba_fees,

            'platform_fee': upload.platform_fee,
            'rembursement_fee': upload.rembursement_fee,

            'expense_chart_img': upload.expense_chart_img,
            'sales_chart_img': upload.sales_chart_img,
            'qtd_pie_chart': upload.qtd_pie_chart,
            'ytd_pie_chart': upload.ytd_pie_chart,

            'total_cous': upload.total_cous,
            'total_amazon_fee': upload.total_amazon_fee,
            'profit_chart_img': upload.profit_chart_img,

            'cm2_profit': upload.cm2_profit,
            'cm2_margins': upload.cm2_margins,
            'acos': upload.acos,

            'rembursment_vs_cm2_margins': upload.rembursment_vs_cm2_margins,
            'advertising_total': upload.advertising_total,
            'reimbursement_vs_sales': upload.reimbursement_vs_sales,
            'taxncredit': upload.taxncredit,
            'unit_sold': upload.unit_sold,

            'otherwplatform': upload.platform_fee,
        })

    return jsonify({'uploads': response}), 200


def resolve_country(country, currency):
    country = (country or "").lower()
    currency = (currency or "").lower()

    # 1. If country = global
    if country == "global":
        if currency == "usd":
            return "global"
        elif currency == "inr":
            return "global_inr"
        elif currency == "gbp":
            return "global_gbp"
        elif currency == "cad":
            return "global_cad"
        else:
            return "global"  # default fallback

    # 2. If country = uk
    if country == "uk":
        if currency == "usd":
            return "uk_usd"
        else:
            return "uk"  # default for all other currencies

    # 3. Default (no special logic)
    return country


@upload_bp.route('/upload_history2', methods=['GET'])
def upload_history2():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Authorization token is missing or invalid'}), 401

    token = auth_header.split(' ')[1]
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=['HS256'])
        user_id = payload['user_id']
    except jwt.ExpiredSignatureError:
        return jsonify({'error': 'Token has expired'}), 401
    except jwt.InvalidTokenError:
        return jsonify({'error': 'Invalid token'}), 401

    range_type = request.args.get('range')
    month = request.args.get('month')
    year = request.args.get('year')
    quarter = request.args.get('quarter')

    country_param = (request.args.get('country', '') or '').lower()

    # ✅ only use homeCurrency for GLOBAL
    if country_param == "global":
        currency_param = (request.args.get('homeCurrency') or 'USD').lower()
    else:
        currency_param = None

    country = resolve_country(country_param, currency_param)

    # Try to infer range_type if not provided
    if not range_type:
        if month and year:
            range_type = 'monthly'
        elif quarter and year:
            range_type = 'quarterly'
        elif year:
            range_type = 'yearly'
        else:
            return jsonify({
                'error': 'Invalid range parameters. Must specify range type or provide appropriate parameters to infer range.'
            }), 400

    try:
        year_num = int(year)
    except (TypeError, ValueError):
        return jsonify({'error': 'Year must be a valid number.'}), 400

    def summarize_uploads(uploads):
        total_sales = sum(upload.total_sales or 0 for upload in uploads)
        total_profit = sum(upload.total_profit or 0 for upload in uploads)
        total_expense = sum(upload.total_expense or 0 for upload in uploads)
        advertising_total = sum(upload.advertising_total or 0 for upload in uploads)
        cm2_profit = sum(upload.cm2_profit or 0 for upload in uploads)
        total_amazon_fee = sum(upload.total_amazon_fee or 0 for upload in uploads)
        total_cous = sum(upload.total_cous or 0 for upload in uploads)
        otherwplatform = sum(upload.platform_fee or 0 for upload in uploads)
        taxncredit = sum(upload.taxncredit or 0 for upload in uploads)
        unit_sold = sum(upload.unit_sold or 0 for upload in uploads)

        return {
            'total_sales': total_sales,
            'total_profit': total_profit,
            'total_expense': total_expense,
            'advertising_total': advertising_total,
            'cm2_profit': cm2_profit,
            'total_amazon_fee': total_amazon_fee,
            'total_cous': total_cous,
            'otherwplatform': otherwplatform,
            'taxncredit': taxncredit,
            'unit_sold': unit_sold,
        }

    # ---------------- comparison helpers ----------------

    month_order = [
        'january', 'february', 'march',
        'april', 'may', 'june',
        'july', 'august', 'september',
        'october', 'november', 'december'
    ]

    quarter_months = {
        'Q1': ['january', 'february', 'march'],
        'Q2': ['april', 'may', 'june'],
        'Q3': ['july', 'august', 'september'],
        'Q4': ['october', 'november', 'december']
    }

    def get_previous_month(m: str, y: int):
        m = (m or '').lower()
        if m not in month_order:
            return None, None
        idx = month_order.index(m)
        if idx == 0:
            return month_order[-1], y - 1
        return month_order[idx - 1], y

    def get_quarter_from_month(m: str):
        m = (m or '').lower()
        for q, months in quarter_months.items():
            if m in months:
                return q
        return None

    def get_previous_quarter(q: str, y: int):
        order = ['Q1', 'Q2', 'Q3', 'Q4']
        if q not in order:
            return None, None
        idx = order.index(q)
        if idx == 0:
            return 'Q4', y - 1
        return order[idx - 1], y

    def fetch_monthly_summary(m: str, y: int):
        if not m or y is None:
            return None
        ups = UploadHistory.query.filter_by(
            user_id=user_id,
            year=y,
            month=m.lower(),
            country=country
        ).all()
        return summarize_uploads(ups) if ups else None

    def fetch_quarterly_summary(q: str, y: int):
        if not q or y is None or q not in quarter_months:
            return None
        ups = UploadHistory.query.filter(
            UploadHistory.user_id == user_id,
            UploadHistory.year == y,
            UploadHistory.month.in_(quarter_months[q]),
            UploadHistory.country == country
        ).all()
        return summarize_uploads(ups) if ups else None

    def fetch_yearly_summary(y: int):
        if y is None:
            return None
        ups = UploadHistory.query.filter_by(
            user_id=user_id,
            year=y,
            country=country
        ).all()
        return summarize_uploads(ups) if ups else None

    # ---------------- main logic ----------------

    if range_type == 'monthly' and month and year:
        month_l = month.lower()

        uploads = UploadHistory.query.filter_by(
            user_id=user_id,
            year=year_num,
            month=month_l,
            country=country
        ).all()

        current_summary = summarize_uploads(uploads)

        # last month (prev month)
        prev_m, prev_y = get_previous_month(month_l, year_num)
        last_month_summary = fetch_monthly_summary(prev_m, prev_y) if prev_m else None

        # last quarter (previous quarter from the month you are viewing)
        current_q = get_quarter_from_month(month_l)
        prev_q, prev_q_y = get_previous_quarter(current_q, year_num) if current_q else (None, None)
        last_quarter_summary = fetch_quarterly_summary(prev_q, prev_q_y) if prev_q else None

        # last year (same month previous year)
        last_year_summary = fetch_monthly_summary(month_l, year_num - 1)

        return jsonify({
            'uploads': [u.id for u in uploads],
            'summary': current_summary,
            'summaryComparisons': {
                'lastMonth': last_month_summary,
                'lastQuarter': last_quarter_summary,
                'lastYear': last_year_summary
            }
        })

    elif range_type == 'quarterly' and quarter and year:
        if quarter not in quarter_months:
            return jsonify({'error': 'Quarter must be one of: Q1, Q2, Q3, Q4'}), 400

        uploads = UploadHistory.query.filter(
            UploadHistory.user_id == user_id,
            UploadHistory.year == year_num,
            UploadHistory.month.in_(quarter_months[quarter]),
            UploadHistory.country == country
        ).all()

        current_summary = summarize_uploads(uploads)

        # last quarter (previous quarter)
        prev_q, prev_q_y = get_previous_quarter(quarter, year_num)
        last_quarter_summary = fetch_quarterly_summary(prev_q, prev_q_y) if prev_q else None

        # last year (same quarter last year)
        last_year_summary = fetch_quarterly_summary(quarter, year_num - 1)

        return jsonify({
            'uploads': [u.id for u in uploads],
            'summary': current_summary,
            'summaryComparisons': {
                'lastMonth': None,
                'lastQuarter': last_quarter_summary,
                'lastYear': last_year_summary
            }
        })

    elif range_type == 'yearly' and year:
        uploads = UploadHistory.query.filter_by(
            user_id=user_id,
            year=year_num,
            country=country
        ).all()

        current_summary = summarize_uploads(uploads)

        # last year
        last_year_summary = fetch_yearly_summary(year_num - 1)

        return jsonify({
            'uploads': [u.id for u in uploads],
            'summary': current_summary,
            'summaryComparisons': {
                'lastMonth': None,
                'lastQuarter': None,
                'lastYear': last_year_summary
            }
        })

    else:
        return jsonify({'error': 'Invalid range parameters'}), 400



@upload_bp.route('/upload_historyforacos', methods=['GET'])
def upload_historyforacos():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Authorization token is missing or invalid'}), 401

    token = auth_header.split(' ')[1]
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=['HS256'])
        user_id = payload['user_id']
    except jwt.ExpiredSignatureError:
        return jsonify({'error': 'Token has expired'}), 401
    except jwt.InvalidTokenError:
        return jsonify({'error': 'Invalid token'}), 401

    # Fetching query params
    country = request.args.get('country')
    month = request.args.get('month')
    year = request.args.get('year')

    # Month name to number mapping
    month_map = {
        'january': 1, 'february': 2, 'march': 3, 'april': 4,
        'may': 5, 'june': 6, 'july': 7, 'august': 8,
        'september': 9, 'october': 10, 'november': 11, 'december': 12
    }

    # Convert month name to number if it's a string
    if month:
        try:
            month_num = int(month)
        except ValueError:
            month_num = month_map.get(month.lower())
            if month_num is None:
                return jsonify({'error': f'Invalid month name: {month}'}), 400
            month = month_num  # Store as integer for filtering

    # Convert year to integer if present
    if year:
        try:
            year = int(year)
        except ValueError:
            return jsonify({'error': 'Invalid year format'}), 400

    # Query with optional filters
    uploads = UploadHistory.query.filter_by(user_id=user_id)
    
    if country:
        uploads = uploads.filter_by(country=country)
    if month:
        uploads = uploads.filter_by(month=month)
    if year:
        uploads = uploads.filter_by(year=year)
    
    uploads = uploads.all()

    # Month number to name mapping
    month_names = {
        1: 'january', 2: 'february', 3: 'march', 4: 'april',
        5: 'may', 6: 'june', 7: 'july', 8: 'august',
        9: 'september', 10: 'october', 11: 'november', 12: 'december'
    }

    response = []
    for upload in uploads:
        # Convert numeric month to month name for display
        month_name = month_names.get(upload.month, str(upload.month))
        
        table_name = f"user_{upload.country}_{month_name}{upload.year}_data"
        response.append({
            'month': month_name,  # Use month name for display
            'month_num': upload.month,  # Keep numeric month for filtering if needed
            'year': upload.year,
            'country': upload.country,
            'file_name': table_name,
            'total_sales': upload.total_sales,
            'total_profit': upload.total_profit,
            'total_expense': upload.total_expense,
            'total_fba_fees': upload.total_fba_fees,
            'platform_fee': upload.platform_fee,
            'rembursement_fee': upload.rembursement_fee,
            'expense_chart_img': upload.expense_chart_img,
            'sales_chart_img': upload.sales_chart_img,
            'profit_chart_img': upload.profit_chart_img,
            'cm2_profit': upload.cm2_profit,
            'cm2_margins': upload.cm2_margins,
            'acos': upload.acos,
            'rembursment_vs_cm2_margins': upload.rembursment_vs_cm2_margins,
            'advertising_total': upload.advertising_total,
            'reimbursement_vs_sales': upload.reimbursement_vs_sales,
            'taxncredit': upload.taxncredit,
        })

    return jsonify({'uploads': response})


